import ctypes

# Substitua 'caminho_para_seu_ícone.ico' pelo caminho real do seu ícone
caminho_icone = 'C:\\Users\\Benedito\\Documents\\GitHub\\Carteirinhas\\CARTEIRA.ico'

# Define o ícone para o arquivo .exe
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")
ctypes.windll.kernel32.SetConsoleIcon(ctypes.windll.shell32.ExtractIconW(0, caminho_icone, 0))


import openpyxl
import os
from tkinter import messagebox, Tk

# Caminhos para as planilhas
pasta_automacoes = 'C:/Users/Benedito/Documents/GitHub/Carteirinhas'
lista_piloto_path = os.path.join(pasta_automacoes, 'ListaPiloto-2023.xlsx')
modelo_carterinha_path = os.path.join(pasta_automacoes, 'ModeloCarteirinha.xlsx')

# Verificar a existência do arquivo ListaPiloto-2023
if not os.path.exists(lista_piloto_path):
    root = Tk()
    root.withdraw()
    messagebox.showerror("Erro", "Certifique-se de que o arquivo ListaPiloto-2023.xlsx existe.")
    root.destroy()
    exit()

# Abrir a planilha Lista Piloto
lista_piloto = openpyxl.load_workbook(lista_piloto_path)

# Iterar sobre todas as planilhas na Lista Piloto
for sheet_turma in lista_piloto.sheetnames:
    # Abre a planilha da turma
    sheet_piloto_turma = lista_piloto[sheet_turma]

    # Obter os nomes da lista
    nomes = [cell.value for row in sheet_piloto_turma.iter_rows(min_row=12, max_row=39, min_col=2, max_col=2) for cell in row]

    # Obter os dados da turma, período e nome da professora
    turma = sheet_piloto_turma['F8'].value
    periodo = sheet_piloto_turma['C8'].value
    nome_professora = sheet_piloto_turma['E10'].value

    # Criar uma nova planilha chamada DADOS
    dados_workbook = openpyxl.Workbook()
    dados_sheet = dados_workbook.active

    # Preencher a nova planilha com as informações
    dados_sheet['A1'] = 'Nomes'
    dados_sheet.append(nomes)
    dados_sheet['A8'] = 'Turma'
    dados_sheet['B8'] = turma
    dados_sheet['A9'] = 'Período'
    dados_sheet['B9'] = periodo
    dados_sheet['A10'] = 'Nome da Professora'
    dados_sheet['B10'] = nome_professora

    # Salvar a nova planilha
    dados_workbook.save(os.path.join(pasta_automacoes, 'Dados.xlsx'))

    # Verificar a existência do arquivo ModeloCarteirinha.xlsx
    if not os.path.exists(modelo_carterinha_path):
        root = Tk()
        root.withdraw()
        messagebox.showerror("Erro", "Certifique-se de que o arquivo ModeloCarteirinha.xlsx existe.")
        root.destroy()
        exit()

    # Abrir planilha Modelo Carteirinha
    modelo_carterinha = openpyxl.load_workbook(modelo_carterinha_path)
    sheet_carterinha = modelo_carterinha.active

    # Inicializar variáveis de posição NOME ALUNO
    linha_inicial_c12 = 12

    # Iterar sobre os dados copiados
    for i in range(0, len(nomes), 4):
        # Atribuir valores às células correspondentes na planilha Modelo Carteirinha
        sheet_carterinha.cell(row=linha_inicial_c12, column=3, value=nomes[i])
        sheet_carterinha.cell(row=linha_inicial_c12, column=10, value=nomes[i + 1])
        sheet_carterinha.cell(row=linha_inicial_c12 + 14, column=3, value=nomes[i + 2])
        sheet_carterinha.cell(row=linha_inicial_c12 + 14, column=10, value=nomes[i + 3])

        # Atualizar posição para próxima colagem
        linha_inicial_c12 += 28

    # Inicializar variáveis de posição TURMA
    linha_inicial_c13 = 13

    # Iterar sobre os dados copiados
    for i in range(0, len(nomes), 4):
        # Atribuir valores às células correspondentes na planilha Modelo Carteirinha
        sheet_carterinha.cell(row=linha_inicial_c13, column=3, value=turma)
        sheet_carterinha.cell(row=linha_inicial_c13, column=10, value=turma)
        sheet_carterinha.cell(row=linha_inicial_c13 + 14, column=3, value=turma)
        sheet_carterinha.cell(row=linha_inicial_c13 + 14, column=10, value=turma)

        # Atualizar posição para próxima colagem
        linha_inicial_c13 += 28

    # Inicializar variáveis de posição PERIODO
    linha_inicial_c14 = 14

    # Iterar sobre os dados copiados
    for i in range(0, len(nomes), 4):
        # Atribuir valores às células correspondentes na planilha Modelo Carteirinha
        sheet_carterinha.cell(row=linha_inicial_c14, column=3, value=periodo)
        sheet_carterinha.cell(row=linha_inicial_c14, column=10, value=periodo)
        sheet_carterinha.cell(row=linha_inicial_c14 + 14, column=3, value=periodo)
        sheet_carterinha.cell(row=linha_inicial_c14 + 14, column=10, value=periodo)

        # Atualizar posição para próxima colagem
        linha_inicial_c14 += 28

    # Inicializar variáveis de posição NOME PROFESSORA
    linha_inicial_e14 = 14

    # Iterar sobre os dados copiados
    for i in range(0, len(nomes), 4):
        # Atribuir valores às células correspondentes na planilha Modelo Carteirinha
        sheet_carterinha.cell(row=linha_inicial_e14, column=4, value='Profº ' + nome_professora)  # Adicionando o prefixo "Profº"
        sheet_carterinha.cell(row=linha_inicial_e14, column=11, value='Profº ' + nome_professora)  # Adicionando o prefixo "Profº"
        sheet_carterinha.cell(row=linha_inicial_e14 + 14, column=4, value='Profº ' + nome_professora)  # Adicionando o prefixo "Profº"
        sheet_carterinha.cell(row=linha_inicial_e14 + 14, column=11, value='Profº ' + nome_professora)  # Adicionando o prefixo "Profº"

        # Atualizar posição para próxima colagem
        linha_inicial_e14 += 28

    # Salvar e fechar planilha Modelo Carteirinha
    nome_turma = turma.replace(" ", "_")  # Substituir espaços por underscores, se necessário
    nome_arquivo_carteirinhas = f'Carteirinhas_{nome_turma}.xlsx'
    modelo_carterinha.save(os.path.join(pasta_automacoes, nome_arquivo_carteirinhas))
    modelo_carterinha.close()

    # Excluir a planilha Dados
    os.remove(os.path.join(pasta_automacoes, 'Dados.xlsx'))

# Fechar planilha Lista Piloto
lista_piloto.close()

# Exibir pop-up informando que as carteirinhas estão prontas
root = Tk()
root.withdraw()  # Esconde a janela principal
messagebox.showinfo('Concluído', 'Todas as carteirinhas foram geradas')

# Finalizar o aplicativo Tkinter
root.destroy()
